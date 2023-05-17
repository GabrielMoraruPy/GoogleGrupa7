import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

urls = [
    f"https://www.mai.gov.ro/informare-covid-19-grupul-de-comunicare-strategica-{i}-martie-ora-13-00-2/" for i in range(1,6)
#     "https://www.mai.gov.ro/informare-covid-19-grupul-de-comunicare-strategica-2-martie-ora-13-00-2/",
#     "https://www.mai.gov.ro/informare-covid-19-grupul-de-comunicare-strategica-3-martie-ora-13-00-2/",
#     "https://www.mai.gov.ro/informare-covid-19-grupul-de-comunicare-strategica-4-martie-ora-13-00-2/",
#     "https://www.mai.gov.ro/informare-covid-19-grupul-de-comunicare-strategica-6-martie-ora-13-00-2/"
]

data = {}
lista_int = []
for url in urls:
    page = requests.get(url).content
    soup = BeautifulSoup(page, "html.parser")

    table = soup.find("table")

    for tr in table.find_all("tr")[1:]:
        cells = tr.find_all("td")

        judet = cells[1].text.strip()
        numar_cazuri = cells[2].text.strip().replace(".","").replace("*","")

        if judet not in data:
            data[judet] = [0] * len(urls)

        if numar_cazuri == "":
            break
        else:
            data[judet][urls.index(url)] = numar_cazuri
            lista_int.append(int(numar_cazuri))

        numar_cazuri_int = int(numar_cazuri)


workbook = openpyxl.Workbook()
sheet = workbook.active

header = ["NR. CRT", "Judet"] + [f"0{day}.03" if day < 10 else f"{day}.03" for day in range(1, len(urls)+1)]
sheet.append(header)

nr_crt = 1
for judet, cazuri in data.items():
    if nr_crt == 45:
        break
    row = [nr_crt, judet] + cazuri
    sheet.append(row)
    nr_crt += 1

print(lista_int)
print(sum(lista_int))

suma_totala = ["Total", ""] + [sum([float(data[judet][i].replace('.', '').replace("*","")) if data[judet][i] else 0 for judet in data]) for i in range(len(urls))]
sheet.append(suma_totala)

font = Font(name='Calibri', size=12, color='FF000000')
border = Border(
    left=Side(border_style='medium', color='FF000000'),
    right=Side(border_style='medium', color='FF000000'),
    top=Side(border_style='medium', color='FF000000'),
    bottom=Side(border_style='medium', color='FF000000')
)
alignment = Alignment(horizontal='center', vertical='center')
fill = PatternFill(fill_type='solid', fgColor='DDDDDDDD')

for cell in sheet[1]:
    cell.font = font
    cell.border = border
    cell.alignment = alignment
    cell.fill = fill

for row in sheet.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = alignment

for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column_letter].width = adjusted_width

workbook.save("date_tema.xlsx")
